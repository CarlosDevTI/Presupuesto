from __future__ import annotations

from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook


def read_workbook_matrix(
    file_source: str | Path | BinaryIO,
    sheet_name: str | None = None,
) -> tuple[str, list[list[object]]]:
    workbook = load_workbook(filename=file_source, data_only=True, read_only=True)
    selected_sheet = sheet_name or workbook.sheetnames[0]
    worksheet = workbook[selected_sheet]
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return selected_sheet, rows
